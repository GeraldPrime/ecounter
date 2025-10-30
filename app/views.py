from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from django.db.models import Sum, Q
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import json
import random
import math

# PDF imports
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from .models import PollingUnit, VoteAllocation, AllocatedResult, UploadSession
from .utils import detect_vote_count_field, validate_vote_count_field
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required

def signin_view(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_staff:
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials or not an admin user.')
    return render(request, 'signin.html')

@login_required
def dashboard(request):
    """Main dashboard view"""
    total_units = PollingUnit.objects.count()
    total_allocations = VoteAllocation.objects.count()
    total_pvc_45 = PollingUnit.objects.aggregate(
        total=Sum('pvc_45_percent')
    )['total'] or 0

    # Calculate average PVC per unit
    average_pvc_per_unit = 0
    if total_units > 0:
        average_pvc_per_unit = total_pvc_45 / total_units

    recent_allocations = VoteAllocation.objects.order_by('-created_at')[:5]
    
    # Get the most recent upload session info
    latest_upload = UploadSession.objects.order_by('-created_at').first()
    current_vote_field = latest_upload.vote_count_field_name if latest_upload else "45% PVC COLLECTION"

    context = {
        'total_units': total_units,
        'total_allocations': total_allocations,
        'total_pvc_45': total_pvc_45,
        'average_pvc_per_unit': average_pvc_per_unit,
        'recent_allocations': recent_allocations,
        'current_vote_field': current_vote_field,
    }
    return render(request, 'vote_allocation/dashboard.html', context)

# def upload_data(request):
#     """Handle Excel file upload"""
#     if request.method == 'POST' and request.FILES.get('excel_file'):
#         excel_file = request.FILES['excel_file']
        
#         try:
#             # Read Excel file
#             df = pd.read_excel(excel_file)
            
#             # Debug: print column names and first few rows
#             print("Excel columns:", df.columns.tolist())
#             print("Data shape:", df.shape)
#             print("First row data:", df.iloc[0].to_dict() if len(df) > 0 else "No data")
            
#             # Clear existing data
#             PollingUnit.objects.all().delete()
            
#             # Import data
#             created_count = 0
#             errors = []
            
#             for index, row in df.iterrows():
#                 try:
#                     # Skip empty rows
#                     if pd.isna(row.get('S/NO')) or row.get('S/NO') == '':
#                         continue
                    
#                     # Create polling unit with just the basic data
#                     polling_unit = PollingUnit.objects.create(
#                         sno=int(float(str(row.get('S/NO', 0)).replace(',', ''))),
#                         state=str(row.get('STATE', '')).strip(),
#                         lga=str(row.get('LGA', '')).strip(),
#                         ra=str(row.get('RA', '')).strip(),
#                         delim=str(row.get('DELIM', '')).strip(),
#                         register_voter_2023=str(row.get('REGISTER VOTER AS AT 2023', '')).strip(),
#                         registered_voter_2024=int(float(str(row.get('REGISTERED VOTER AS AT 2024', 0)).replace(',', ''))),
#                         pvc_collected=int(float(str(row.get('NO OF PVC COLLECTED ', 0)).replace(',', ''))),
#                         balance_uncollected=int(float(str(row.get('BALANCE OF UNCOLECTED PVCs', 0)).replace(',', ''))),
#                         pvc_45_percent=float(str(row.get('45% PVC COLLECTION', 0)).replace(',', '')),
#                         # Set party fields to 0 for now - they'll be calculated during allocation
#                         aa_original=0,
#                         ad_original=0,
#                         adc_original=0,
#                         apc_original=0,
#                         lp_original=0,
#                         pdp_original=0,
#                     )
#                     created_count += 1
                    
#                     # Print progress every 100 records
#                     if created_count % 100 == 0:
#                         print(f"Imported {created_count} records...")
                        
#                 except Exception as row_error:
#                     error_msg = f"Error in row {index + 2}: {str(row_error)}"
#                     print(error_msg)
#                     errors.append(error_msg)
#                     continue
            
#             if created_count > 0:
#                 messages.success(request, f'Successfully imported {created_count} polling units.')
#                 if errors:
#                     messages.warning(request, f'Encountered {len(errors)} errors during import.')
#             else:
#                 messages.error(request, 'No valid data was imported. Please check your Excel file format.')
            
#             return redirect('dashboard')
            
#         except Exception as e:
#             error_msg = f'Error importing data: {str(e)}'
#             print(error_msg)
#             messages.error(request, error_msg)

#     elif request.method == 'POST':
#         messages.error(request, 'No file was selected. Please choose an Excel file.')
    
#     return render(request, 'vote_allocation/upload.html')


@login_required
def upload_data(request):
    """Handle Excel file upload with dynamic field detection"""
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            df = pd.read_excel(excel_file)
            
            # Detect vote count field
            detected_field = detect_vote_count_field(df.columns.tolist())
            
            if detected_field:
                # Validate the detected field
                is_valid, validation_msg = validate_vote_count_field(df, detected_field)
                
                if is_valid:
                    # Proceed with import using detected field
                    return process_excel_import(request, df, detected_field)
                else:
                    # Field detected but invalid, show selection interface
                    return show_field_selection(request, df, validation_msg)
            else:
                # No field detected, show selection interface
                return show_field_selection(request, df, "No vote count field automatically detected")
            
        except Exception as e:
            error_msg = f'Error reading Excel file: {str(e)}'
            print(error_msg)
            messages.error(request, error_msg)
    
    elif request.method == 'POST' and request.POST.get('vote_count_field'):
        # User selected a field manually
        excel_file = request.FILES.get('excel_file')
        if excel_file:
            try:
                df = pd.read_excel(excel_file)
                selected_field = request.POST.get('vote_count_field')
                
                # Validate selected field
                is_valid, validation_msg = validate_vote_count_field(df, selected_field)
                
                if is_valid:
                    return process_excel_import(request, df, selected_field)
                else:
                    messages.error(request, f'Selected field is invalid: {validation_msg}')
                    return show_field_selection(request, df, validation_msg)
            except Exception as e:
                messages.error(request, f'Error processing file: {str(e)}')
    
    elif request.method == 'POST':
        messages.error(request, 'No file was selected. Please choose an Excel file.')

    return render(request, 'vote_allocation/upload.html')

def show_field_selection(request, df, error_msg=None):
    """Show field selection interface when automatic detection fails"""
    context = {
        'columns': df.columns.tolist(),
        'error_msg': error_msg,
        'sample_data': df.head(3).to_dict('records') if len(df) > 0 else []
    }
    return render(request, 'vote_allocation/field_selection.html', context)

def process_excel_import(request, df, vote_count_field):
    """Process Excel import with the specified vote count field"""
    try:
        # Clear existing data
        PollingUnit.objects.all().delete()
        AllocatedResult.objects.all().delete()
        
        created_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                if pd.isna(row.get('S/NO')) or row.get('S/NO') == '':
                    continue
                
                # Get vote count from the specified field
                vote_count_raw = float(str(row.get(vote_count_field, 0)).replace(',', ''))
                vote_count_rounded = round(vote_count_raw)
                
                polling_unit = PollingUnit.objects.create(
                    sno=int(float(str(row.get('S/NO', 0)).replace(',', ''))),
                    state=str(row.get('STATE', '')).strip(),
                    lga=str(row.get('LGA', '')).strip(),
                    ra=str(row.get('RA', '')).strip(),
                    delim=str(row.get('DELIM', '')).strip(),
                    register_voter_2023=str(row.get('REGISTER VOTER AS AT 2023', '')).strip(),
                    registered_voter_2024=int(float(str(row.get('REGISTERED VOTER AS AT 2024', 0)).replace(',', ''))),
                    pvc_collected=int(float(str(row.get('NO OF PVC COLLECTED ', 0)).replace(',', ''))),
                    balance_uncollected=int(float(str(row.get('BALANCE OF UNCOLECTED PVCs', 0)).replace(',', ''))),
                    pvc_45_percent=vote_count_rounded,  # Store the vote count in pvc_45_percent field
                    aa_original=0,
                    ad_original=0,
                    adc_original=0,
                    apc_original=0,
                    lp_original=0,
                    pdp_original=0,
                    nrm_original=0,
                    nnpp_original=0,
                    prp_original=0,
                    sdp_original=0,
                    ypp_original=0,
                    yp_original=0,
                    zlp_original=0,
                    a_original=0,
                    aac_original=0,
                    adp_original=0,
                    apm_original=0,
                    apga_original=0,
                    app_original=0,
                    bp_original=0,
                )
                created_count += 1
                
                if created_count % 100 == 0:
                    print(f"Imported {created_count} records...")
                    
            except Exception as row_error:
                error_msg = f"Error in row {index + 2}: {str(row_error)}"
                print(error_msg)
                errors.append(error_msg)
                continue
        
        # Store upload session info
        UploadSession.objects.create(
            vote_count_field_name=vote_count_field,
            total_records=created_count
        )
        
        if created_count > 0:
            messages.success(request, f'Successfully imported {created_count} polling units using field "{vote_count_field}".')
            if errors:
                messages.warning(request, f'Encountered {len(errors)} errors during import.')
        else:
            messages.error(request, 'No valid data was imported. Please check your Excel file format.')
        
        return redirect('dashboard')
        
    except Exception as e:
        error_msg = f'Error importing data: {str(e)}'
        print(error_msg)
        messages.error(request, error_msg)
        return redirect('upload_data')

@login_required
def polling_units_list(request):
    """List all polling units with pagination"""
    units = PollingUnit.objects.all()

    # Search functionality
    search = request.GET.get('search')
    if search:
        units = units.filter(
            Q(state__icontains=search) |
            Q(lga__icontains=search) |
            Q(delim__icontains=search)
        )

    paginator = Paginator(units, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search': search,
    }
    return render(request, 'vote_allocation/polling_units.html', context)

# def create_allocation(request):
#     """Create new vote allocation"""
#     if request.method == 'POST':
#         try:
#             allocation = VoteAllocation.objects.create(
#                 name=request.POST['name'],
#                 description=request.POST.get('description', ''),
#                 aa_percentage=float(request.POST.get('aa_percentage', 0)),
#                 ad_percentage=float(request.POST.get('ad_percentage', 0)),
#                 adc_percentage=float(request.POST.get('adc_percentage', 0)),
#                 apc_percentage=float(request.POST.get('apc_percentage', 0)),
#                 lp_percentage=float(request.POST.get('lp_percentage', 0)),
#                 pdp_percentage=float(request.POST.get('pdp_percentage', 0)),
#             )
            
#             if not allocation.is_valid_allocation():
#                 messages.warning(request, 
#                     f'Total percentage is {allocation.total_percentage():.1f}%. Should be 100%.')
            
#             # Calculate allocated results
#             calculate_allocated_results(allocation)

#             messages.success(request, 'Vote allocation created successfully!')
#             return redirect('view_allocation_results', allocation_id=allocation.id)

#         except Exception as e:
#             messages.error(request, f'Error creating allocation: {str(e)}')

#     return render(request, 'vote_allocation/create_allocation.html')


# def create_allocation(request):
#     """Create new vote allocation"""
#     if request.method == 'POST':
#         try:
#             # Check if we have polling units
#             units_count = PollingUnit.objects.count()
#             if units_count == 0:
#                 messages.error(request, 'No polling units found. Please upload data first.')
#                 return redirect('upload_data')
            
#             allocation = VoteAllocation.objects.create(
#                 name=request.POST['name'],
#                 description=request.POST.get('description', ''),
#                 aa_percentage=float(request.POST.get('aa_percentage', 0)),
#                 ad_percentage=float(request.POST.get('ad_percentage', 0)),
#                 adc_percentage=float(request.POST.get('adc_percentage', 0)),
#                 apc_percentage=float(request.POST.get('apc_percentage', 0)),
#                 lp_percentage=float(request.POST.get('lp_percentage', 0)),
#                 pdp_percentage=float(request.POST.get('pdp_percentage', 0)),
#             )
            
#             if not allocation.is_valid_allocation():
#                 messages.warning(request, 
#                     f'Total percentage is {allocation.total_percentage():.1f}%. Should be 100%.')
            
#             # SIMPLE RELIABLE CALCULATION - Create results for all polling units
#             polling_units = PollingUnit.objects.all()
#             results = []
            
#             for unit in polling_units:
#                 if unit.pvc_collected > 0:  # Only process units with votes
#                     # Simple calculation without randomization
#                     aa_votes = int(unit.pvc_collected * (allocation.aa_percentage / 100))
#                     ad_votes = int(unit.pvc_collected * (allocation.ad_percentage / 100))
#                     adc_votes = int(unit.pvc_collected * (allocation.adc_percentage / 100))
#                     apc_votes = int(unit.pvc_collected * (allocation.apc_percentage / 100))
#                     lp_votes = int(unit.pvc_collected * (allocation.lp_percentage / 100))
#                     pdp_votes = int(unit.pvc_collected * (allocation.pdp_percentage / 100))
#                     total_votes = aa_votes + ad_votes + adc_votes + apc_votes + lp_votes + pdp_votes
                    
#                     results.append(AllocatedResult(
#                         polling_unit=unit,
#                         vote_allocation=allocation,
#                         aa_votes=aa_votes,
#                         ad_votes=ad_votes,
#                         adc_votes=adc_votes,
#                         apc_votes=apc_votes,
#                         lp_votes=lp_votes,
#                         pdp_votes=pdp_votes,
#                         total_votes=total_votes,
#                     ))
            
#             # Create all results at once
#             AllocatedResult.objects.bulk_create(results)
            
#             messages.success(request, f'Vote allocation created successfully! Generated {len(results)} results.')
#             return redirect('view_allocation_results', allocation_id=allocation.id)
            
#         except Exception as e:
#             messages.error(request, f'Error creating allocation: {str(e)}')

#     return render(request, 'vote_allocation/create_allocation.html')


def create_allocation(request):
    """Create new vote allocation"""
    if request.method == 'POST':
        try:
            units_count = PollingUnit.objects.count()
            if units_count == 0:
                messages.error(request, 'No polling units found. Please upload data first.')
                return redirect('upload_data')
            
            allocation = VoteAllocation.objects.create(
                name=request.POST['name'],
                description=request.POST.get('description', ''),
                aa_percentage=float(request.POST.get('aa_percentage', 0)),
                ad_percentage=float(request.POST.get('ad_percentage', 0)),
                adc_percentage=float(request.POST.get('adc_percentage', 0)),
                apc_percentage=float(request.POST.get('apc_percentage', 0)),
                lp_percentage=float(request.POST.get('lp_percentage', 0)),
                pdp_percentage=float(request.POST.get('pdp_percentage', 0)),
                nrm_percentage=float(request.POST.get('nrm_percentage', 0)),
                nnpp_percentage=float(request.POST.get('nnpp_percentage', 0)),
                prp_percentage=float(request.POST.get('prp_percentage', 0)),
                sdp_percentage=float(request.POST.get('sdp_percentage', 0)),
                ypp_percentage=float(request.POST.get('ypp_percentage', 0)),
                yp_percentage=float(request.POST.get('yp_percentage', 0)),
                zlp_percentage=float(request.POST.get('zlp_percentage', 0)),
                a_percentage=float(request.POST.get('a_percentage', 0)),
                aac_percentage=float(request.POST.get('aac_percentage', 0)),
                adp_percentage=float(request.POST.get('adp_percentage', 0)),
                apm_percentage=float(request.POST.get('apm_percentage', 0)),
                apga_percentage=float(request.POST.get('apga_percentage', 0)),
                app_percentage=float(request.POST.get('app_percentage', 0)),
                bp_percentage=float(request.POST.get('bp_percentage', 0)),
            )
            
            if not allocation.is_valid_allocation():
                messages.warning(request, 
                    f'Total percentage is {allocation.total_percentage():.1f}%. Should be 100%.')
            
            # Use 45% PVC instead of PVC collected
            polling_units = PollingUnit.objects.all()
            results = []
            
            for unit in polling_units:
                if unit.pvc_45_percent > 0:  # Use 45% PVC instead of pvc_collected
                    base_votes = unit.pvc_45_percent  # Changed from pvc_collected
                    
                    aa_votes = int(base_votes * (allocation.aa_percentage / 100))
                    ad_votes = int(base_votes * (allocation.ad_percentage / 100))
                    adc_votes = int(base_votes * (allocation.adc_percentage / 100))
                    apc_votes = int(base_votes * (allocation.apc_percentage / 100))
                    lp_votes = int(base_votes * (allocation.lp_percentage / 100))
                    pdp_votes = int(base_votes * (allocation.pdp_percentage / 100))
                    nrm_votes = int(base_votes * (allocation.nrm_percentage / 100))
                    nnpp_votes = int(base_votes * (allocation.nnpp_percentage / 100))
                    prp_votes = int(base_votes * (allocation.prp_percentage / 100))
                    sdp_votes = int(base_votes * (allocation.sdp_percentage / 100))
                    ypp_votes = int(base_votes * (allocation.ypp_percentage / 100))
                    yp_votes = int(base_votes * (allocation.yp_percentage / 100))
                    zlp_votes = int(base_votes * (allocation.zlp_percentage / 100))
                    a_votes = int(base_votes * (allocation.a_percentage / 100))
                    aac_votes = int(base_votes * (allocation.aac_percentage / 100))
                    adp_votes = int(base_votes * (allocation.adp_percentage / 100))
                    apm_votes = int(base_votes * (allocation.apm_percentage / 100))
                    apga_votes = int(base_votes * (allocation.apga_percentage / 100))
                    app_votes = int(base_votes * (allocation.app_percentage / 100))
                    bp_votes = int(base_votes * (allocation.bp_percentage / 100))
                    total_votes = (aa_votes + ad_votes + adc_votes + apc_votes + lp_votes + pdp_votes +
                                 nrm_votes + nnpp_votes + prp_votes + sdp_votes + ypp_votes + yp_votes +
                                 zlp_votes + a_votes + aac_votes + adp_votes + apm_votes + apga_votes + app_votes + bp_votes)
                    
                    results.append(AllocatedResult(
                        polling_unit=unit,
                        vote_allocation=allocation,
                        aa_votes=aa_votes,
                        ad_votes=ad_votes,
                        adc_votes=adc_votes,
                        apc_votes=apc_votes,
                        lp_votes=lp_votes,
                        pdp_votes=pdp_votes,
                        nrm_votes=nrm_votes,
                        nnpp_votes=nnpp_votes,
                        prp_votes=prp_votes,
                        sdp_votes=sdp_votes,
                        ypp_votes=ypp_votes,
                        yp_votes=yp_votes,
                        zlp_votes=zlp_votes,
                        a_votes=a_votes,
                        aac_votes=aac_votes,
                        adp_votes=adp_votes,
                        apm_votes=apm_votes,
                        apga_votes=apga_votes,
                        app_votes=app_votes,
                        bp_votes=bp_votes,
                        total_votes=total_votes,
                    ))
            
            AllocatedResult.objects.bulk_create(results)
            
            messages.success(request, f'Vote allocation created successfully! Generated {len(results)} results.')
            return redirect('view_allocation_results', allocation_id=allocation.id)
            
        except Exception as e:
            messages.error(request, f'Error creating allocation: {str(e)}')

    return render(request, 'vote_allocation/create_allocation.html')


@login_required
def allocations_list(request):
    """List all vote allocations"""
    allocations = VoteAllocation.objects.order_by('-created_at')

    context = {
        'allocations': allocations,
    }
    return render(request, 'vote_allocation/allocations_list.html', context)

def calculate_allocated_results(allocation):
    """Calculate and save realistic allocated results for all polling units"""
    # Clear existing results for this allocation
    AllocatedResult.objects.filter(vote_allocation=allocation).delete()
    
    # Get all polling units
    polling_units = PollingUnit.objects.all()
    
    # Calculate for each unit
    results = []
    for unit in polling_units:
        # Use the actual PVC collected as the base for vote distribution
        base_votes = unit.pvc_collected
        
        # Add some realism - not everyone votes, typically 70-95% turnout
        turnout_rate = random.uniform(0.75, 0.95)
        actual_votes = int(base_votes * turnout_rate)
        
        # Calculate target votes for each party based on allocation percentages
        target_aa = (allocation.aa_percentage / 100) * actual_votes
        target_ad = (allocation.ad_percentage / 100) * actual_votes
        target_adc = (allocation.adc_percentage / 100) * actual_votes
        target_apc = (allocation.apc_percentage / 100) * actual_votes
        target_lp = (allocation.lp_percentage / 100) * actual_votes
        target_pdp = (allocation.pdp_percentage / 100) * actual_votes
        
        # Add realistic variation (±5-15% from target to simulate real voting patterns)
        def add_realistic_variation(target_votes, variation_range=0.10):
            if target_votes == 0:
                return 0
            variation = random.uniform(-variation_range, variation_range)
            result = target_votes * (1 + variation)
            return max(0, int(round(result)))
        
        # Calculate actual votes with variation
        aa_votes = add_realistic_variation(target_aa)
        ad_votes = add_realistic_variation(target_ad)
        adc_votes = add_realistic_variation(target_adc)
        apc_votes = add_realistic_variation(target_apc)
        lp_votes = add_realistic_variation(target_lp)
        pdp_votes = add_realistic_variation(target_pdp)
        
        # Calculate total and adjust if necessary to match actual votes
        calculated_total = aa_votes + ad_votes + adc_votes + apc_votes + lp_votes + pdp_votes
        
        # Adjust the largest party's votes to match the actual vote count
        if calculated_total != actual_votes:
            difference = actual_votes - calculated_total
            # Find the party with the highest allocation to adjust
            party_votes = [
                ('aa', aa_votes), ('ad', ad_votes), ('adc', adc_votes),
                ('apc', apc_votes), ('lp', lp_votes), ('pdp', pdp_votes)
            ]
            party_votes.sort(key=lambda x: x[1], reverse=True)
            
            # Adjust the largest party
            if party_votes[0][0] == 'aa':
                aa_votes += difference
            elif party_votes[0][0] == 'ad':
                ad_votes += difference
            elif party_votes[0][0] == 'adc':
                adc_votes += difference
            elif party_votes[0][0] == 'apc':
                apc_votes += difference
            elif party_votes[0][0] == 'lp':
                lp_votes += difference
            elif party_votes[0][0] == 'pdp':
                pdp_votes += difference
        
        # Ensure no negative votes
        aa_votes = max(0, aa_votes)
        ad_votes = max(0, ad_votes)
        adc_votes = max(0, adc_votes)
        apc_votes = max(0, apc_votes)
        lp_votes = max(0, lp_votes)
        pdp_votes = max(0, pdp_votes)
        
        total_votes = aa_votes + ad_votes + adc_votes + apc_votes + lp_votes + pdp_votes
        
        results.append(AllocatedResult(
            polling_unit=unit,
            vote_allocation=allocation,
            aa_votes=aa_votes,
            ad_votes=ad_votes,
            adc_votes=adc_votes,
            apc_votes=apc_votes,
            lp_votes=lp_votes,
            pdp_votes=pdp_votes,
            total_votes=total_votes,
        ))
    
    # Bulk create results
    AllocatedResult.objects.bulk_create(results)
    print(f"Created realistic allocation results for {len(results)} polling units")

# FIXED - Single view_allocation_results function
@login_required
def view_allocation_results(request, allocation_id):
    """View allocation details and results with party percentages displayed"""
    allocation = get_object_or_404(VoteAllocation, id=allocation_id)
    results = AllocatedResult.objects.filter(vote_allocation=allocation).select_related('polling_unit')
    
    # Pagination for results
    paginator = Paginator(results, 50)  # Show more results per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate totals and verify percentages
    totals = results.aggregate(
        total_aa=Sum('aa_votes'),
        total_ad=Sum('ad_votes'),
        total_adc=Sum('adc_votes'),
        total_apc=Sum('apc_votes'),
        total_lp=Sum('lp_votes'),
        total_pdp=Sum('pdp_votes'),
        total_nrm=Sum('nrm_votes'),
        total_nnpp=Sum('nnpp_votes'),
        total_prp=Sum('prp_votes'),
        total_sdp=Sum('sdp_votes'),
        total_ypp=Sum('ypp_votes'),
        total_yp=Sum('yp_votes'),
        total_zlp=Sum('zlp_votes'),
        total_a=Sum('a_votes'),
        total_aac=Sum('aac_votes'),
        total_adp=Sum('adp_votes'),
        total_apm=Sum('apm_votes'),
        total_apga=Sum('apga_votes'),
        total_app=Sum('app_votes'),
        total_bp=Sum('bp_votes'),
        grand_total=Sum('total_votes'),
    )
    
    # Calculate actual percentages achieved
    grand_total = totals['grand_total'] or 1  # Avoid division by zero
    actual_percentages = {
        'aa': (totals['total_aa'] or 0) / grand_total * 100,
        'ad': (totals['total_ad'] or 0) / grand_total * 100,
        'adc': (totals['total_adc'] or 0) / grand_total * 100,
        'apc': (totals['total_apc'] or 0) / grand_total * 100,
        'lp': (totals['total_lp'] or 0) / grand_total * 100,
        'pdp': (totals['total_pdp'] or 0) / grand_total * 100,
        'nrm': (totals['total_nrm'] or 0) / grand_total * 100,
        'nnpp': (totals['total_nnpp'] or 0) / grand_total * 100,
        'prp': (totals['total_prp'] or 0) / grand_total * 100,
        'sdp': (totals['total_sdp'] or 0) / grand_total * 100,
        'ypp': (totals['total_ypp'] or 0) / grand_total * 100,
        'yp': (totals['total_yp'] or 0) / grand_total * 100,
        'zlp': (totals['total_zlp'] or 0) / grand_total * 100,
        'a': (totals['total_a'] or 0) / grand_total * 100,
        'aac': (totals['total_aac'] or 0) / grand_total * 100,
        'adp': (totals['total_adp'] or 0) / grand_total * 100,
        'apm': (totals['total_apm'] or 0) / grand_total * 100,
        'apga': (totals['total_apga'] or 0) / grand_total * 100,
        'app': (totals['total_app'] or 0) / grand_total * 100,
        'bp': (totals['total_bp'] or 0) / grand_total * 100,
    }

    context = {
        'allocation': allocation,
        'page_obj': page_obj,
        'totals': totals,
        'actual_percentages': actual_percentages,
    }
    return render(request, 'vote_allocation/view_allocation_results.html', context)

# NEW - Full data view like polling units but with party allocations
@login_required
def view_allocation_full_data(request, allocation_id):
    """View all allocated results in a table format like polling units"""
    allocation = get_object_or_404(VoteAllocation, id=allocation_id)
    results = AllocatedResult.objects.filter(vote_allocation=allocation).select_related('polling_unit')

    # Search functionality
    search = request.GET.get('search')
    if search:
        results = results.filter(
            Q(polling_unit__state__icontains=search) |
            Q(polling_unit__lga__icontains=search) |
            Q(polling_unit__delim__icontains=search)
        )

    paginator = Paginator(results, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Calculate totals
    totals = results.aggregate(
        total_aa=Sum('aa_votes'),
        total_ad=Sum('ad_votes'),
        total_adc=Sum('adc_votes'),
        total_apc=Sum('apc_votes'),
        total_lp=Sum('lp_votes'),
        total_pdp=Sum('pdp_votes'),
        total_nrm=Sum('nrm_votes'),
        total_nnpp=Sum('nnpp_votes'),
        total_prp=Sum('prp_votes'),
        total_sdp=Sum('sdp_votes'),
        total_ypp=Sum('ypp_votes'),
        total_yp=Sum('yp_votes'),
        total_zlp=Sum('zlp_votes'),
        total_a=Sum('a_votes'),
        total_aac=Sum('aac_votes'),
        total_adp=Sum('adp_votes'),
        total_apm=Sum('apm_votes'),
        total_apga=Sum('apga_votes'),
        total_app=Sum('app_votes'),
        total_bp=Sum('bp_votes'),
        grand_total=Sum('total_votes'),
    )
    
    context = {
        'allocation': allocation,
        'page_obj': page_obj,
        'search': search,
        'totals': totals,
    }
    return render(request, 'vote_allocation/allocation_full_data.html', context)

@login_required
def download_allocation_excel(request, allocation_id):
    """Download allocation results as Excel file with complete totals"""
    allocation = get_object_or_404(VoteAllocation, id=allocation_id)
    results = AllocatedResult.objects.filter(vote_allocation=allocation).select_related('polling_unit').order_by('polling_unit__sno')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vote Allocation Results"
    
    # Get the current vote field name
    latest_upload = UploadSession.objects.order_by('-created_at').first()
    vote_field_name = latest_upload.vote_count_field_name if latest_upload else "45% PVC COLLECTION"
    
    headers = [
        'S/NO', 'STATE', 'LGA', 'RA', 'DELIM', 'REGISTER VOTER AS AT 2023',
        'REGISTERED VOTER AS AT 2024', 'NO OF PVC COLLECTED', 'BALANCE OF UNCOLLECTED PVCs',
        vote_field_name,
        f'AA ({allocation.aa_percentage}%)', 
        f'AD ({allocation.ad_percentage}%)', 
        f'ADC ({allocation.adc_percentage}%)', 
        f'APC ({allocation.apc_percentage}%)', 
        f'LP ({allocation.lp_percentage}%)', 
        f'PDP ({allocation.pdp_percentage}%)', 
        f'NRM ({allocation.nrm_percentage}%)',
        f'NNPP ({allocation.nnpp_percentage}%)',
        f'PRP ({allocation.prp_percentage}%)',
        f'SDP ({allocation.sdp_percentage}%)',
        f'YPP ({allocation.ypp_percentage}%)',
        f'YP ({allocation.yp_percentage}%)',
        f'ZLP ({allocation.zlp_percentage}%)',
        f'A ({allocation.a_percentage}%)',
        f'AAC ({allocation.aac_percentage}%)',
        f'APM ({allocation.apm_percentage}%)',
        f'APGA ({allocation.apga_percentage}%)',
        f'APP ({allocation.app_percentage}%)',
        f'BP ({allocation.bp_percentage}%)',
        'Invalid Votes',
        'TOTAL'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_num, result in enumerate(results, 2):
        unit = result.polling_unit
        ws.cell(row=row_num, column=1, value=unit.sno)
        ws.cell(row=row_num, column=2, value=unit.state)
        ws.cell(row=row_num, column=3, value=unit.lga)
        ws.cell(row=row_num, column=4, value=unit.ra)
        ws.cell(row=row_num, column=5, value=unit.delim)
        ws.cell(row=row_num, column=6, value=unit.register_voter_2023)
        ws.cell(row=row_num, column=7, value=unit.registered_voter_2024)
        ws.cell(row=row_num, column=8, value=unit.pvc_collected)
        ws.cell(row=row_num, column=9, value=unit.balance_uncollected)
        ws.cell(row=row_num, column=10, value=int(unit.pvc_45_percent))  # Show as whole number
        ws.cell(row=row_num, column=11, value=result.aa_votes)
        ws.cell(row=row_num, column=12, value=result.ad_votes)
        ws.cell(row=row_num, column=13, value=result.adc_votes)
        ws.cell(row=row_num, column=14, value=result.apc_votes)
        ws.cell(row=row_num, column=15, value=result.lp_votes)
        ws.cell(row=row_num, column=16, value=result.pdp_votes)
        ws.cell(row=row_num, column=17, value=result.nrm_votes)
        ws.cell(row=row_num, column=18, value=result.nnpp_votes)
        ws.cell(row=row_num, column=19, value=result.prp_votes)
        ws.cell(row=row_num, column=20, value=result.sdp_votes)
        ws.cell(row=row_num, column=21, value=result.ypp_votes)
        ws.cell(row=row_num, column=22, value=result.yp_votes)
        ws.cell(row=row_num, column=23, value=result.zlp_votes)
        ws.cell(row=row_num, column=24, value=result.a_votes)
        ws.cell(row=row_num, column=25, value=result.aac_votes)
        ws.cell(row=row_num, column=26, value=result.apm_votes)
        ws.cell(row=row_num, column=27, value=result.apga_votes)
        ws.cell(row=row_num, column=28, value=result.app_votes)
        ws.cell(row=row_num, column=29, value=result.bp_votes)
        invalid_votes = max(int(unit.pvc_45_percent) - int(result.total_votes), 0)
        ws.cell(row=row_num, column=30, value=invalid_votes)
        ws.cell(row=row_num, column=31, value=result.total_votes)
    
    # TOTALS ROW - Calculate totals for ALL numeric columns
    total_row = len(results) + 2
    ws.cell(row=total_row, column=1, value="TOTALS")
    
    # Calculate column totals
    total_reg_2024 = sum(unit.registered_voter_2024 for unit in PollingUnit.objects.filter(id__in=[r.polling_unit.id for r in results]))
    total_pvc_collected = sum(unit.pvc_collected for unit in PollingUnit.objects.filter(id__in=[r.polling_unit.id for r in results]))
    total_balance = sum(unit.balance_uncollected for unit in PollingUnit.objects.filter(id__in=[r.polling_unit.id for r in results]))
    total_pvc_45 = sum(unit.pvc_45_percent for unit in PollingUnit.objects.filter(id__in=[r.polling_unit.id for r in results]))
    
    vote_totals = results.aggregate(
        total_aa=Sum('aa_votes'),
        total_ad=Sum('ad_votes'),
        total_adc=Sum('adc_votes'),
        total_apc=Sum('apc_votes'),
        total_lp=Sum('lp_votes'),
        total_pdp=Sum('pdp_votes'),
        total_nrm=Sum('nrm_votes'),
        total_nnpp=Sum('nnpp_votes'),
        total_prp=Sum('prp_votes'),
        total_sdp=Sum('sdp_votes'),
        total_ypp=Sum('ypp_votes'),
        total_yp=Sum('yp_votes'),
        total_zlp=Sum('zlp_votes'),
        total_a=Sum('a_votes'),
        total_aac=Sum('aac_votes'),
        total_apm=Sum('apm_votes'),
        total_apga=Sum('apga_votes'),
        total_app=Sum('app_votes'),
        total_bp=Sum('bp_votes'),
        grand_total=Sum('total_votes'),
    )
    
    # Insert totals
    ws.cell(row=total_row, column=7, value=total_reg_2024)
    ws.cell(row=total_row, column=8, value=total_pvc_collected)
    ws.cell(row=total_row, column=9, value=total_balance)
    ws.cell(row=total_row, column=10, value=int(total_pvc_45))
    ws.cell(row=total_row, column=11, value=vote_totals['total_aa'])
    ws.cell(row=total_row, column=12, value=vote_totals['total_ad'])
    ws.cell(row=total_row, column=13, value=vote_totals['total_adc'])
    ws.cell(row=total_row, column=14, value=vote_totals['total_apc'])
    ws.cell(row=total_row, column=15, value=vote_totals['total_lp'])
    ws.cell(row=total_row, column=16, value=vote_totals['total_pdp'])
    ws.cell(row=total_row, column=17, value=vote_totals['total_nrm'])
    ws.cell(row=total_row, column=18, value=vote_totals['total_nnpp'])
    ws.cell(row=total_row, column=19, value=vote_totals['total_prp'])
    ws.cell(row=total_row, column=20, value=vote_totals['total_sdp'])
    ws.cell(row=total_row, column=21, value=vote_totals['total_ypp'])
    ws.cell(row=total_row, column=22, value=vote_totals['total_yp'])
    ws.cell(row=total_row, column=23, value=vote_totals['total_zlp'])
    ws.cell(row=total_row, column=24, value=vote_totals['total_a'])
    ws.cell(row=total_row, column=25, value=vote_totals['total_aac'])
    ws.cell(row=total_row, column=26, value=vote_totals['total_apm'])
    ws.cell(row=total_row, column=27, value=vote_totals['total_apga'])
    ws.cell(row=total_row, column=28, value=vote_totals['total_app'])
    ws.cell(row=total_row, column=29, value=vote_totals['total_bp'])
    total_invalid = max(int(total_pvc_45) - int(vote_totals['grand_total'] or 0), 0)
    ws.cell(row=total_row, column=30, value=total_invalid)
    ws.cell(row=total_row, column=31, value=vote_totals['grand_total'])
    
    # Style totals row
    for col in range(1, 32):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"vote_allocation_{allocation.name.replace(' ', '_')}_{allocation.id}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


# NEW - PDF Download Function
@login_required
def download_allocation_pdf(request, allocation_id):
    """Download allocation results as PDF file"""
    allocation = get_object_or_404(VoteAllocation, id=allocation_id)
    results = AllocatedResult.objects.filter(vote_allocation=allocation).select_related('polling_unit').order_by('polling_unit__sno')[:100]  # Limit for PDF
    
    # Create PDF
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=12,
        alignment=1  # Center alignment
    )
    
    # Story container
    story = []
    
    # Title
    story.append(Paragraph(f"Vote Allocation Results: {allocation.name}", title_style))
    story.append(Spacer(1, 12))
    
    # Allocation details
    allocation_text = f"""
    <b>Allocation Percentages:</b><br/>
    AA: {allocation.aa_percentage}% | AD: {allocation.ad_percentage}% | ADC: {allocation.adc_percentage}%<br/>
    APC: {allocation.apc_percentage}% | LP: {allocation.lp_percentage}% | PDP: {allocation.pdp_percentage}%<br/>
    <b>Created:</b> {allocation.created_at.strftime('%Y-%m-%d %H:%M')}
    """
    story.append(Paragraph(allocation_text, styles['Normal']))
    story.append(Spacer(1, 12))
    
    # Table data
    # Get the current vote field name
    latest_upload = UploadSession.objects.order_by('-created_at').first()
    vote_field_name = latest_upload.vote_count_field_name if latest_upload else "45% PVC COLLECTION"
    
    table_data = [
        ['S/NO', 'State', 'LGA', 'Polling Unit', vote_field_name[:10], 'AA', 'AD', 'ADC', 'APC', 'LP', 'PDP', 'NRM', 'NNPP', 'PRP', 'SDP', 'YPP', 'YP', 'ZLP', 'A', 'AAC', 'APM', 'APGA', 'APP', 'BP', 'Total']
    ]
    
    for result in results:
        unit = result.polling_unit
        table_data.append([
            str(unit.sno),
            unit.state[:10],  # Truncate for PDF
            unit.lga[:10],
            unit.delim[:20],
            str(unit.pvc_collected),
            str(result.aa_votes),
            str(result.ad_votes),
            str(result.adc_votes),
            str(result.apc_votes),
            str(result.lp_votes),
            str(result.pdp_votes),
            str(result.nrm_votes),
            str(result.nnpp_votes),
            str(result.prp_votes),
            str(result.sdp_votes),
            str(result.ypp_votes),
            str(result.yp_votes),
            str(result.zlp_votes),
            str(result.a_votes),
            str(result.aac_votes),
            str(result.apm_votes),
            str(result.apga_votes),
            str(result.app_votes),
            str(result.bp_votes),
            str(result.total_votes)
        ])
    
    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    
    # Calculate totals
    totals = results.aggregate(
        total_aa=Sum('aa_votes'),
        total_ad=Sum('ad_votes'),
        total_adc=Sum('adc_votes'),
        total_apc=Sum('apc_votes'),
        total_lp=Sum('lp_votes'),
        total_pdp=Sum('pdp_votes'),
        total_nrm=Sum('nrm_votes'),
        total_nnpp=Sum('nnpp_votes'),
        total_prp=Sum('prp_votes'),
        total_sdp=Sum('sdp_votes'),
        total_ypp=Sum('ypp_votes'),
        total_yp=Sum('yp_votes'),
        total_zlp=Sum('zlp_votes'),
        total_a=Sum('a_votes'),
        total_aac=Sum('aac_votes'),
        total_apm=Sum('apm_votes'),
        total_apga=Sum('apga_votes'),
        total_app=Sum('app_votes'),
        total_bp=Sum('bp_votes'),
        grand_total=Sum('total_votes'),
    )
    
    # Add totals
    story.append(Spacer(1, 12))
    totals_text = f"""
    <b>TOTALS:</b><br/>
    AA: {totals['total_aa']} | AD: {totals['total_ad']} | ADC: {totals['total_adc']} | APC: {totals['total_apc']}<br/>
    LP: {totals['total_lp']} | PDP: {totals['total_pdp']} | NRM: {totals['total_nrm']} | NNPP: {totals['total_nnpp']}<br/>
    PRP: {totals['total_prp']} | SDP: {totals['total_sdp']} | YPP: {totals['total_ypp']} | YP: {totals['total_yp']}<br/>
    ZLP: {totals['total_zlp']} | A: {totals['total_a']} | AAC: {totals['total_aac']} | APM: {totals['total_apm']}<br/>
    APGA: {totals['total_apga']} | APP: {totals['total_app']} | BP: {totals['total_bp']}<br/>
    <b>Grand Total: {totals['grand_total']}</b>
    """
    story.append(Paragraph(totals_text, styles['Normal']))
    
    if len(results) == 100:
        story.append(Spacer(1, 12))
        story.append(Paragraph("<i>Note: Only first 100 records shown in PDF. Download Excel for complete data.</i>", styles['Normal']))
    
    # Build PDF
    doc.build(story)
    
    # Return response
    output.seek(0)
    filename = f"vote_allocation_{allocation.name.replace(' ', '_')}_{allocation.id}.pdf"
    response = HttpResponse(output.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@csrf_exempt
@login_required
def validate_allocation(request):
    """AJAX endpoint to validate allocation percentages"""
    if request.method == 'POST':
        data = json.loads(request.body)
        
        total = (
            float(data.get('aa_percentage', 0)) +
            float(data.get('ad_percentage', 0)) +
            float(data.get('adc_percentage', 0)) +
            float(data.get('apc_percentage', 0)) +
            float(data.get('lp_percentage', 0)) +
            float(data.get('pdp_percentage', 0)) +
            float(data.get('nrm_percentage', 0)) +
            float(data.get('nnpp_percentage', 0)) +
            float(data.get('prp_percentage', 0)) +
            float(data.get('sdp_percentage', 0)) +
            float(data.get('ypp_percentage', 0)) +
            float(data.get('yp_percentage', 0)) +
            float(data.get('zlp_percentage', 0)) +
            float(data.get('a_percentage', 0)) +
            float(data.get('aac_percentage', 0)) +
            float(data.get('adp_percentage', 0)) +
            float(data.get('apm_percentage', 0)) +
            float(data.get('apga_percentage', 0)) +
            float(data.get('app_percentage', 0)) +
            float(data.get('bp_percentage', 0))
        )
        
        return JsonResponse({
            'total': round(total, 2),
            'is_valid': abs(total - 100.0) < 0.01
        })
    
    return JsonResponse({'error': 'Invalid request'})