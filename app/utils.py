# utils.py - Helper functions
import pandas as pd
from io import BytesIO
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

# Field mapping for vote count detection
VOTE_COUNT_FIELD_MAPPING = [
    "45% PVC COLLECTION",
    "ACCREDITATION", 
    "VOTES CAST",
    "TOTAL VOTES",
    "VOTER TURNOUT",
    "ACTUAL VOTES",
    "VOTES RECORDED",
    "BALLOT CAST",
    "VOTING NUMBERS",
    "VOTES",
    "TURNOUT",
    "ACCREDITED VOTERS",
    "VOTERS ACCREDITED",
    "NUMBER OF VOTERS",
    "VOTE COUNT",
    "TOTAL ACCREDITATION"
]

def detect_vote_count_field(df_columns):
    """
    Detect the vote count field from Excel columns using multiple strategies
    Returns the field name if found, None if not found
    """
    if not df_columns:
        return None
    
    # Convert all column names to uppercase for case-insensitive matching
    columns_upper = [col.upper().strip() for col in df_columns]
    
    # Strategy 1: Look for exact matches first
    for field_name in VOTE_COUNT_FIELD_MAPPING:
        if field_name.upper() in columns_upper:
            # Find the original case version
            for col in df_columns:
                if col.upper().strip() == field_name.upper():
                    return col
    
    # Strategy 2: Look for partial matches (contains keywords)
    vote_keywords = ['VOTE', 'ACCREDITATION', 'CAST', 'BALLOT', 'TURN', 'ACCREDITED']
    for col in df_columns:
        col_upper = col.upper().strip()
        if any(keyword in col_upper for keyword in vote_keywords):
            return col
    
    # Strategy 3: Look for numeric columns that might contain vote counts
    numeric_candidates = []
    for col in df_columns:
        col_upper = col.upper().strip()
        # Skip known non-vote columns
        skip_terms = ['S/NO', 'STATE', 'LGA', 'RA', 'DELIM', 'REGISTER', 'PVC', 'BALANCE']
        if not any(term in col_upper for term in skip_terms):
            # Check if column name suggests it's a count
            if any(term in col_upper for term in ['NUMBER', 'COUNT', 'TOTAL', 'NO', 'NUM']):
                numeric_candidates.append(col)
    
    if numeric_candidates:
        return numeric_candidates[0]  # Return first candidate
    
    return None

def validate_vote_count_field(df, field_name):
    """
    Validate that the selected field contains valid numeric vote count data
    """
    if not field_name or field_name not in df.columns:
        return False, "Field not found in Excel file"
    
    try:
        # Convert to numeric, handling commas and other formatting
        vote_data = df[field_name].astype(str).str.replace(',', '').str.replace(' ', '')
        vote_data = pd.to_numeric(vote_data, errors='coerce')
        
        # Check if we have valid numeric data
        valid_count = vote_data.notna().sum()
        total_count = len(vote_data)
        
        if valid_count == 0:
            return False, "Field contains no valid numeric data"
        
        if valid_count / total_count < 0.5:  # Less than 50% valid data
            return False, f"Field contains mostly invalid data ({valid_count}/{total_count} valid)"
        
        # Check if values are reasonable (positive numbers)
        if vote_data.min() < 0:
            return False, "Field contains negative values"
        
        return True, f"Field validated successfully ({valid_count}/{total_count} valid records)"
        
    except Exception as e:
        return False, f"Error validating field: {str(e)}"

def validate_excel_columns(df):
    """Validate that the Excel file has required columns (excluding vote count field)"""
    required_columns = [
        'S/NO', 'STATE', 'LGA', 'RA', 'DELIM',
        'REGISTER VOTER AS AT 2023', 'REGISTERED VOTER AS AT 2024',
        'NO OF PVC COLLECTED ', 'BALANCE OF UNCOLECTED PVCs'
    ]
    
    missing_columns = []
    for col in required_columns:
        if col not in df.columns:
            missing_columns.append(col)
    
    return missing_columns

def clean_excel_data(df):
    """Clean and prepare Excel data for import"""
    # Fill NaN values
    df = df.fillna(0)
    
    # Convert numeric columns
    numeric_cols = [
        'REGISTERED VOTER AS AT 2024', 'NO OF PVC COLLECTED ',
        'BALANCE OF UNCOLECTED PVCs', '45% PVC COLLECTION',
        'AA', 'AD', 'ADC', 'APC', 'LP', 'PDP'
    ]
    
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
    
    return df

def export_allocation_to_excel(allocation, results):
    """Export allocation results to Excel with formatting"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vote Allocation Results"
    
    # Headers
    headers = [
        'S/NO', 'STATE', 'LGA', 'RA', 'DELIM', 'REGISTER VOTER AS AT 2023',
        'REGISTERED VOTER AS AT 2024', 'NO OF PVC COLLECTED', 'BALANCE OF UNCOLLECTED PVCs',
        '45% PVC COLLECTION', 'AA', 'AD', 'ADC', 'APC', 'LP', 'PDP', 'TOTAL'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Add allocation info as a comment or separate sheet
    ws.cell(row=1, column=len(headers) + 2, value=f"Allocation: {allocation.name}")
    ws.cell(row=2, column=len(headers) + 2, value=f"APC: {allocation.apc_percentage}%")
    ws.cell(row=3, column=len(headers) + 2, value=f"LP: {allocation.lp_percentage}%")
    ws.cell(row=4, column=len(headers) + 2, value=f"PDP: {allocation.pdp_percentage}%")
    
    return wb

#